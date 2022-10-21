import sqlalchemy 
import pandas as pd
import os
from datetime import datetime
from datetime import date
import json
import logging
import re
from bs4 import BeautifulSoup
import requests as req
import pickle
import platform
import spacy
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from collections import Counter
import seaborn as sns
import pandas as pd
import numpy as np
'''
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait 
from selenium.webdriver.support import expected_conditions as EC
'''

class Model:
    def __init__(self, conf, urls):
        if urls:
            self.urls = urls

        if not conf:
            file_name = [f for f in os.listdir('{}'.format("\\".join(__file__.split("\\")[:-1]))) if os.path.isfile(os.path.join('{}'.format("\\".join(__file__.split("\\")[:-1])), f)) if re.search(r'\w+\.json', f) is not None] 
            if len(file_name) == 1:
                file_path = os.path.join('{}'.format("\\".join(__file__.split("\\")[:-1])), file_name[0])
                self.config = json.load(open(f'{file_path}'))
            else:
                print('Error: More that one config-file found')
                logging.critical("Error: More that one config-file found!")
                quit()
        else:
            self.config = json.load(open(conf))

        if platform.system() == 'Windows':
            self.file_path = os.path.realpath(__file__).split('\\') # for windows 
        else:
            self.file_path = os.path.realpath(__file__).split('/') #for mac
        self.file_path = '/'.join([self.file_path[i] for i in range(len(self.file_path)-1)])

        name = 'log' + str(date.today()) + '.log'

        logging.basicConfig(filename=os.path.join(os.path.join(self.file_path, 'log'), name),format='%(asctime)s %(levelname)-8s %(message)s', datefmt='%Y-%m-%d %H:%M:%S', filemode='w',level=logging.INFO)

class DBConnection(Model):
    def __init__(self, conf = None):
        super().__init__(conf, None)
        logging.info('DB-Connector started.')
        self.engine = None

    def connect(self):
        #Connect to server & create engine
        print('Connecting...')
        logging.info(f"Trying to connect to Server: {self.config['server']}; Database: {self.config['database']}")
        try:
            #con_str = 'Driver={' +  self.config['driver'] + '};''Server=' + self.config['server'] + ';DATABASE=' + self.config['database'] + ';''UID=' + self.config['user'] + ';''PWD=' + self.config['password'] + ';''Trusted_Connection=no;'
            con_str = 'Driver=' +  self.config['driver'] + ';''Server=' + self.config['protocol'] + ':' +  self.config['server'] + '\\' + self.config["instance"] + ';DATABASE=' + self.config['database'] + ';''UID=' + self.config['user'] + ';''PWD=' + self.config['password']
            engine = sqlalchemy.create_engine('mssql+pyodbc:///?odbc_connect={}'.format(con_str), fast_executemany=True, encoding="utf8") #Creating the sqlalchemy engine
            test_flag = pd.read_sql_query('SELECT 1', engine)
            
            print('success!')
            logging.info(f"Successfully connected!")
            self.engine =  engine
        
        except Exception as e:
            logging.warning(e)
            logging.critical('There was a problem with source DB connection!')
            print("Connection failed!")
            raise
    
    def getTableData(self):
        table_name = self.config['table']

        if isinstance(self.engine, None):
            raise ValueError(      
                "Expected encountered connection-error\n"
                f"DBConnection.connect returned engine-object of type None."
            )

        with self.engine.connect() as conn:
            for url in conn.execute(sqlalchemy.text('Select * from [{}].[{}].[{}]'.format(self.config['database'], self.config['schema'], table_name))).fetchall(): 
                self.urls = url

                return url

class Webscrapper(Model):
    def __init__(self, urls = None, conf = None):
        super().__init__(conf, urls)
        self.articles = []
        #self.options = Options()
        #self.options.headless = True
        #self.options.add_argument("--window-size=1920,1200")
        #self.driver = webdriver.Chrome(options=self.options,executable_path=os.path.join(os.path.join(self.file_path, "driver"),"chromedriver.exe")) 
        #self.driver.implicitly_wait(10) 
        #set path to chrome driver
        logging.info('Web-Scrapper started.')
        #logging.info('Using Chrome-drivers: ', os.path.join(os.path.join(self.file_path, "driver"),"chromedriver.exe"))
        #logging.info('headless: ', self.options.headless)

    def getText(self):
        print(f'Starting scrapper for {len(self.urls)} webistes')
        len_sentence = 0
        for link in self.urls:
            temp = req.get(link).text
            #self.driver.get(link)
            #WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//button[text()='Accept All']"))).click()
            #temp = self.driver.page_source
            doc = BeautifulSoup(temp, 'html.parser')
            title =  [item.get_text() for item in doc.find_all(['h1','h2','h3'])]
            title = [t for t in title if t]
            text = [item.get_text() for item in doc.find_all('p')]
            text = ' '.join(text).replace('\n', ' ')
            text = re.split('(?<=[.!?]) +', text)
            len_sentence += len(text)

            if len(text) > 5: self.articles.append([{'url': link, 'data':[[{'title':t} for t in title], [{'text':t} for t in text]]}])
            else: 
                logging.info(f'Could not scrap any text for: {link}')
                print(f'Could not scrap any text for: {link}')

        
        logging.info(f'Scrabbed {len(self.urls)} with {len_sentence} sentences')
        #self.exportText()
        #self.driver.quit()

    def exportText(self):
        print('Saving data to data-file...')
        path = os.path.join(self.file_path, 'Text_Data')
        file_list = [f for f in os.listdir(path) if os.path.isfile(os.path.join(path, f))]
        data = []
        
        for file in file_list:
            if re.search(fr'(data)(.*)',file):
                data.append(re.search(fr'(data)(.*)', file).group())

        if data:
            data = [a.split('.') for a in data]
            data = int(max([a[0][-1] for a in data if a[0][-1].isnumeric()])) + 1
        else:
            data = 1
        
        data_name = f'data{data}'.split('.')[0] + '.p'

        with open(os.path.join(os.path.join(self.file_path, 'Text_Data'), data_name), 'wb') as fp:
            pickle.dump(self.articles, fp, protocol=pickle.HIGHEST_PROTOCOL)

        print('done.')
        logging.info(f"Saved data-file to: {os.path.join(os.path.join(self.file_path, 'Text_Data'), data_name)}")

    def loadText(self, name):
        print(f'loading data from data-file: {name}')
        data_name = name 

        with open(os.path.join(os.path.join(self.file_path, 'Text_Data'), data_name), 'rb') as fp:
            self.articles = pickle.load(fp)

        print('done.')
        logging.info(f"loaded data-file from: {os.path.join(os.path.join(self.file_path, 'Text_Data'), data_name)}")

class NLP(Model):
    def __init__(self, urls = None, name = None, loadName = None, conf = None, loadData = False, saveData = False, savePred = False, loadPred = False, showfreq = False):
        if savePred and loadPred:
            logging.critical('savePred and loadPred can not be true at the same time!')
            raise RuntimeError('savePred and loadPred can not be true at the same time!')

        if saveData and loadData:
            logging.critical('saveData and loadData can not be true at the same time!')
            raise RuntimeError('saveData and loadData can not be true at the same time!')

        if loadPred and not loadName or loadName and not loadPred:
            logging.critical('If you want to load predictions please provide loadName and set loadPred = True')
            raise RuntimeError('If you want to load predictions please provide loadName and set loadPred = True')

        if loadData and not name or name and not loadData:
            logging.critical('If you want to load Data please provide name and set loadData = True')
            raise RuntimeError('If you want to load Data please provide name and set loadData = True')

        if loadPred and urls:
            logging.critical('If loadPred = true, urls can not be provided at the same time!')
            raise RuntimeError('If loadPred = true, urls can not be provided at the same time!')

        if loadData and urls:
            logging.critical('If loadData = true, urls can not be provided at the same time!')
            raise RuntimeError('If loadData = true, urls can not be provided at the same time!')

        super().__init__(conf, urls)
        logging.info('NLP started')

        self.savePred = savePred
        self.loadPred = loadPred
        self.name = name
        self.showfreq = showfreq
        self.cleanedText = []

        self.nlp = spacy.load("en_core_web_trf")
        scrapper = Webscrapper(urls = urls, conf = conf)
        self.summary = []
        self.analyzed = None

        if loadData: scrapper.loadText(name)
        else: scrapper.getText()

        self.articles = scrapper.articles

        self.cleanPred = []

        if saveData:
            scrapper.exportText()

        if not urls:
            self.urls = []
            for article in self.articles:
                for elements in article:
                    self.urls.append(elements['url'])
        
        self.threshold = self.config['threshold']
   
    @staticmethod
    def cleanUpText(articles):
        cleanedText = []

        for article in articles:
            temp = []
            for data in article[0]['data']:
                for i in range(len(data)):
                    clean_sent = re.sub(r'[^\w\s\-\_\'\§\$\%\&\=\#\<\>]','',list(data[i].values())[0])
                    temp.append(clean_sent)
            cleanedText.append(temp)
        
        return cleanedText

    @staticmethod
    def uniqueDict(texts):
        wordSet = []
        for article in texts:
            set_of_words = {}
            for sentence in article:
                set_of_words = set(set_of_words).union(set(sentence))
            wordSet.append(set_of_words)
        return wordSet

    @staticmethod
    def createDicts(articles, pred, wordSet):
        wordDicts = []
        for i in range(len(articles)):
            temp = []
            for j in range(len(articles[i])):
                wordDict = dict.fromkeys(wordSet[i], 0)
                if pred[i][j]: #check if empty
                    for word in pred[i][j]:
                        wordDict[word] += articles[i][j].count(word)
                temp.append([wordDict, j]) #j = index sentence

            wordDicts.append(temp)

        return wordDicts

    @staticmethod
    def createVector(wordDicts):
        vectors = []
        for article in wordDicts:
            temp = []
            for sentence in article:
                temp.append([np.fromiter(sentence[0].values(), dtype=float), sentence[1]])
            
            vectors.append(temp)
        return vectors

    @staticmethod
    def cosineDistance(word_vectors):
        cosDist = []
        for article in word_vectors:
            temp = []
            for i in range(len(article)):
                avg = 0
                for j in range(len(article)):
                    if not np.linalg.norm(article[i][0]) or not np.linalg.norm(article[i][0])*np.linalg.norm(article[j][0]):
                        cosine_similarity = 0
                    else:
                        cosine_similarity = np.dot(article[i][0], article[j][0])/(np.linalg.norm(article[i][0])*np.linalg.norm(article[j][0]))

                    avg += cosine_similarity   
                avg /= len(article)        
                temp.append([avg, article[i][1]])
            cosDist.append(temp)

        return cosDist

    @staticmethod
    def summarize_cos(cosDist, cleanedText):
        summary = [] 
        for i in range(len(cosDist)):
            temp = []
            num_sent = int(np.ceil(len(cosDist[i])**.5)) #number of words for the summary
            sort = sorted(cosDist[i], key=lambda x:x[0], reverse=True)
            text = sorted(sort[:num_sent], key=lambda x:x[1])
            for j in range(num_sent):
                temp.append({'sentence': cleanedText[i][text[j][1]], 'score':text[j][0]})
            
            summary.append(temp)
        
        return summary

    def predict(self):
        predictions = []
        self.predictions = None

        num_pred = 0

        for article in self.articles:
            entities = []
            temp2 = []
            for elements in article:
                for element in elements['data']:
                    for el in element:
                        temp = []
                        doc = self.nlp(list(el.values())[0])

                        for ent in doc.ents:
                            num_pred += 1
                            entities.append({'original' : list(el.values())[0], 'word' : ent.text, 'label' : ent.label_})
                            temp.append(ent.text)
                        
                        temp2.append(temp)
            self.cleanPred.append(temp2)
            predictions.append(entities)

        predictions = [{'url':self.urls[i], 'data' : predictions[i]} for i in range(len(self.articles))]  

        self.predictions = predictions

        logging.info(f'analyzing {len(self.urls)} websites')
        logging.info(f'Made {num_pred} predictions')

        return predictions

    def summarize_freq(self):
        count = []
        texts = []
        pred = []

        for i in range(len(self.predictions)):
            temp = []
            temp2 = []
            temp3 = []
            temp4 = []
            articles = []

            for w in self.predictions[i]['data']:
                temp.append(w['word'])

            temp3.append(temp)
            temp = Counter(temp).most_common(1)
            count.append(temp)

            for sentence in self.articles[i][0]['data']:
                for sen in sentence:
                    temp2 = [*list(sen.values())]
                    temp4 = re.sub('\s+',' ',temp2[0])
                    temp4 = temp4.replace("\n", " ")
                    temp4 = temp4.replace("\t", " ")
                    articles.append(temp4)

            texts.append(articles)
            pred.append(temp3)

        scored = []
        count = [x[0] for x in count]
        pred = [[list(d) for d in c] for c in[list(b.items()) for b in [Counter(a) for a in [x[0] for x in pred]]]]

        for i in range(len(pred)):
            temp = []

            for j in range(len(texts[i])):
                score = 0

                clean_sent = re.sub(r'[^\w\s\-\_\']','',texts[i][j])
                clean_sent = clean_sent.split(' ')

                c = count[i][1]

                for k in range(len(pred[i])):
                    norm = pred[i][k][1]/c #normalizing the scores
                    #ocurence = sum(1 for _ in re.finditer(r'\b%s\b' % re.escape(pred[i][j][0]), texts[i][j]))
                    ocurence = texts[i][j].count(pred[i][j][0])
                    score += ocurence*norm
                
                temp.append({'sentence':texts[i][j], 'score':score})

            scored.append(temp)
        
        for i in range(len(scored)):
            temp = []
            temp2 = []
            s = []
            num_sent = int(np.ceil(len(scored[i])**.5)) #number of words for the summary
            sort = sorted(scored[i], key=lambda x:x['score'], reverse=True)
            for j in range(num_sent):
                temp2.append(sort[j]['sentence'])
                temp.append({'sentence': sort[j]['sentence'], 'score':sort[j]['score']})
            
            self.summary.append(temp)
            s.append(temp2)

        return s, texts

    def foward(self): 
        if not self.loadPred:
            self.predict()
        else:
            self.load_pred(self.name)
        if not self.loadPred and self.savePred:
            self.save_pred()

        if self.showfreq:
            Visualizer.showFrequency(self.predictions)

        self.analyze()
        #self.summarize_freq()
        self.cleanedText = NLP.cleanUpText(self.articles)
        wordSet = NLP.uniqueDict(self.cleanPred)
        wordDicts = NLP.createDicts(self.cleanedText, self.cleanPred, wordSet)
        word_vectors = NLP.createVector(wordDicts)
        cosDist = NLP.cosineDistance(word_vectors)
        self.summary = NLP.summarize_cos(cosDist, self.cleanedText)

        self.exportResults()

    def save_pred(self):
        print('Saving data to pred-file...')
        path = os.path.join(self.file_path, 'Pred_Data')
        file_list = [f for f in os.listdir(path) if os.path.isfile(os.path.join(path, f))]
        data = []
        
        for file in file_list:
            if re.search(fr'(pred)(.*)',file):
                data.append(re.search(fr'(pred)(.*)', file).group())

        if data:
            data = [a.split('.') for a in data]
            data = int(max([a[0][-1] for a in data if a[0][-1].isnumeric()])) + 1
        else:
            data = 1
        
        data_name = f'pred{data}'.split('.')[0] + '.p'

        with open(os.path.join(os.path.join(self.file_path, 'Pred_Data'), data_name), 'wb') as fp:
            pickle.dump(self.articles, fp, protocol=pickle.HIGHEST_PROTOCOL)

        print('done.')
        logging.info(f"Saved pred-file to: {os.path.join(os.path.join(self.file_path, 'Pred_Data'), data_name)}")

    def load_pred(self, name):
        print(f'loading data from pred-file: {name}')
        data_name = name + '.p'

        with open(os.path.join(os.path.join(self.file_path, 'Pred_Data'), data_name), 'rb') as fp:
            file = pickle.load(fp)
        
        logging.info(f"loaded pred-file from: {os.path.join(os.path.join(self.file_path, 'Pred_Data'), data_name)}")
        print('done.')

        return file

    def analyze(self, file = False, name = None):
        if file:
            self.predictions = self.load(name)
        
        unique = [[pred['word'] for pred in prediction['data']] for prediction in self.predictions]
        num = [Counter(s) for s in unique]

        analyzed = []
        for i in range(len(self.predictions)):
            url = self.predictions[i]['url']
            temp = []
            for preds in self.predictions[i]['data']:
                temp.append({'original': preds['original'], 'word':preds['word'], 'label':preds['label'], 'occurence':num[i][preds['word']]})
            
            analyzed.append({'url':url, 'data':temp})
        
        self.analyzed = analyzed

    def exportResults(self):
        print('writing data to a excel-file...')
        logging.info('Writing data to a excel-file')
        workbook = Workbook()
        sheet = workbook.active
        
        sheet["A1"] = "Summarized Sentences"
        sheet["A1"].fill = PatternFill("solid", start_color="4287f5")
        sheet["A1"].font = Font(bold=True)
        sheet["B1"] = "Score"
        sheet["B1"].fill = PatternFill("solid", start_color="38d9d1")
        sheet["B1"].font = Font(bold=True)
        sheet["C1"] = "Word"
        sheet["C1"].fill = PatternFill("solid", start_color="FF6600")
        sheet["C1"].font = Font(bold=True)
        sheet["D1"] = "Label"
        sheet["D1"].fill = PatternFill("solid", start_color="99CC00")
        sheet["D1"].font = Font(bold=True)
        sheet["E1"] = "Occurence"
        sheet["E1"].fill = PatternFill("solid", start_color="339966")
        sheet["E1"].font = Font(bold=True)
        sheet["F1"] = "Meaning of Labels"
        sheet["F1"].fill = PatternFill("solid", start_color="FF0000")
        sheet["F1"].font = Font(bold=True)
        sheet["G1"] = "Sentence"
        sheet["G1"].font = Font(bold=True)
        sheet["H1"] = "Url"
        sheet["H1"].font = Font(bold=True)
        sheet["I1"] = "Timestamp"
        sheet["I1"].font = Font(bold=True)
        
        k = 0
        m = 0
        for i in range(len(self.analyzed)):
            for j in range(len(self.analyzed[i]['data'])):
                if j < len(self.summary[i]):
                    sheet.cell(row=k+2+m, column=1,value=self.summary[i][j]['sentence']) #summarized sentence
                    sheet.cell(row=k+2+m, column=2,value=self.summary[i][j]['score']) #score

                sheet.cell(row=k+2+m, column=3,value=self.analyzed[i]['data'][j]['word']) #Word
                sheet.cell(row=k+2+m, column=4,value=self.analyzed[i]['data'][j]['label']) #Label
                sheet.cell(row=k+2+m, column=5,value=self.analyzed[i]['data'][j]['occurence']) #Occurence
                sheet.cell(row=k+2+m, column=7,value=self.analyzed[i]['data'][j]['original']) #Sentence
                sheet.cell(row=k+2+m, column=8,value=self.analyzed[i]['url']) #url
                sheet.cell(row=k+2+m, column=9,value=datetime.now()) #timestamp
            
                k += 1

                '''
                if j == len(self.analyzed[i]['data']) - 1:
                    m += 1
                '''

        
        #Legend of Labels:
        sheet.cell(row=2, column=6,value="PERSON: People, including fictional")
        sheet.cell(row=3, column=6,value="NORP: Nationalities or religious or political groups")
        sheet.cell(row=4, column=6,value="FACILITY: Buildings, airports, highways, bridges etc.")
        sheet.cell(row=5, column=6,value="ORG: Companies, agencies, institutions, etc.")
        sheet.cell(row=6, column=6,value="GPE: Countries, cities, states")
        sheet.cell(row=7, column=6,value="LOC: Non-GPE locations, mountain ranges, bodies of water")
        sheet.cell(row=8, column=6,value="PRODUCT: Objects, vehicles, foods, etc. (Not services)")
        sheet.cell(row=9, column=6,value="EVENT: Named hurricanes, battles, wars, sports events, etc.")
        sheet.cell(row=10, column=6,value="WORK_OF_ART: Titles of books, songs, etc.")
        sheet.cell(row=11, column=6,value="LAW: Named documents made into laws")
        sheet.cell(row=12, column=6,value="LANGUAGE: Any named language")
        sheet.cell(row=13, column=6,value="DATE: Absolute or relative dates or periods")
        sheet.cell(row=14, column=6,value="TIME: Times smaller than a day")
        sheet.cell(row=15, column=6,value="PERCENT: Percentage, including '%'")
        sheet.cell(row=16, column=6,value="MONEY: Monetary values, including unit")
        sheet.cell(row=17, column=6,value="QUANTITY: Measurements, as of weight or distance")
        sheet.cell(row=18, column=6,value="ORDINAL: 'first', 'second', etc.")
        sheet.cell(row=19, column=6,value="CARDINAL: Numerals that do not fall under another type")

        name = self.config['file_name']+'.xlsx'

        workbook.save(os.path.join(self.file_path, name))
        logging.info(f"Saved excel-file to: {os.path.join(self.file_path, name)}")
        print('done.')

class UrlLoader:
    @staticmethod
    def getFilePath():
        if platform.system() == 'Windows':
            file_path = os.path.realpath(__file__).split('\\') # for windows 
        else:
            file_path = os.path.realpath(__file__).split('/') #for mac
        file_path = '/'.join([file_path[i] for i in range(len(file_path)-1)])

        return file_path

    @staticmethod
    def loadFromFile(name):
        print(f'loading urls from url-file: {name}')

        file_path = UrlLoader.getFilePath()
        file_path = os.path.join(file_path, 'url')

        with open(f"{os.path.join(file_path, name)}",'r',encoding = 'utf-8') as f:
            url = [line.replace("\n", "") for line in f]

        print('done.')
        logging.info(f"loaded urls from: {os.path.join(file_path, name)}")
        
        return url

    @staticmethod
    def saveToFile(urls):
        file_path = UrlLoader.getFilePath()
        file_path = os.path.join(file_path, 'url')
        file_list = [f for f in os.listdir(file_path) if os.path.isfile(os.path.join(file_path, f))]

        data = []
        
        for file in file_list:
            if re.search(fr'(urls)(.*)', file):
                data.append(re.search(fr'(urls)(.*)', file).group())

        if len(data) > 0:
            data = [a.split('.') for a in data]
            data = int(max([a[0][-1] for a in data if a[0][-1].isnumeric()])) + 1
        else:
            data = 1
        
        data_name = f'urls{data}'.split('.')[0] + '.data'

        with open(f"{os.path.join(file_path, data_name)}",'w',encoding = 'utf-8') as f:
            for url in urls:
                f.write(url+'\n')
                    
class Visualizer():
    @staticmethod
    def showFrequency(data):
        sns.set()
        file_path = UrlLoader.getFilePath()
        file_path = os.path.join(file_path, 'visuals')

        print(f"Saving {len(data)} visuals to: {file_path}")

        for i in range(len(data)):
            temp = []

            for w in data[i]['data']:
                temp.append((w['word'], w['label']))
            
            temp = Counter(temp)

            df = pd.DataFrame(temp.items(), columns=['word', 'frequency']).sort_values(by='frequency', ascending=True)
            fig = sns.barplot(x='frequency', y='word', data=df).get_figure()
            fig.savefig(os.path.join(file_path, f"out{i}.png"), bbox_inches='tight') 

        print('done.')
        logging.info(f"Saved {len(data)} visuals to: {file_path}")
